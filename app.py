import streamlit as st
import pandas as pd
import requests
import re
from requests.auth import HTTPBasicAuth
import plotly.express as px
from streamlit_autorefresh import st_autorefresh
from datetime import datetime, timedelta
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors as rl_colors
from reportlab.lib.pagesizes import letter
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import urllib3
import threading
import json

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

session = requests.Session()
retries = Retry(total=5, backoff_factor=3, status_forcelist=[500, 502, 503, 504], raise_on_status=False)
session.mount("https://", HTTPAdapter(max_retries=retries))
session.mount("http://",  HTTPAdapter(max_retries=retries))

st.set_page_config(layout="wide", page_title="QATrackPro", page_icon="🚀")

st.markdown("""
<style>
.block-container { padding-top: 0.8rem !important; padding-bottom: 1rem !important; }
header { display: none !important; }
#MainMenu { visibility: hidden; }
footer { visibility: hidden; }
.section-header {
    font-size: 20px; font-weight: 700;
    padding: 8px 0 4px 0;
    border-bottom: 2px solid #2C3E50;
    margin-bottom: 16px; color: white;
}
.plan-badge {
    display: inline-block; background: #1a3a5c;
    border: 1px solid #3498DB; color: #3498DB;
    border-radius: 6px; padding: 2px 10px;
    font-size: 12px; font-weight: 600; margin: 2px 4px 2px 0;
}
.notif-badge {
    display: inline-block; background: #1a4a2c;
    border: 1px solid #2ECC71; color: #2ECC71;
    border-radius: 6px; padding: 3px 10px;
    font-size: 12px; font-weight: 600;
}
</style>
""", unsafe_allow_html=True)

# ----------------------------------
# 🔐 LOGIN + PASSWORD CHANGE
# ----------------------------------
if "users" not in st.session_state:
    st.session_state.users = {"team1": "123", "team2": "123"}
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if not st.session_state.logged_in:
    _, col_m, _ = st.columns([1, 1.2, 1])
    with col_m:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown("## 🚀 QATrackPro")
        tab_login, tab_pwd = st.tabs(["🔐 Login", "🔑 Change Password"])
        with tab_login:
            u = st.text_input("Username", key="login_user")
            p = st.text_input("Password", type="password", key="login_pass")
            if st.button("Login", use_container_width=True, type="primary"):
                if u in st.session_state.users and st.session_state.users[u] == p:
                    st.session_state.logged_in = True
                    st.session_state.username  = u
                    st.rerun()
                else:
                    st.error("Invalid credentials ❌")
        with tab_pwd:
            u2      = st.text_input("Username",         key="pwd_user")
            old_pwd = st.text_input("Current Password", type="password", key="old_pwd")
            new_pwd = st.text_input("New Password",     type="password", key="new_pwd")
            cfm_pwd = st.text_input("Confirm Password", type="password", key="cfm_pwd")
            if st.button("Change Password", use_container_width=True):
                if u2 not in st.session_state.users:
                    st.error("User not found ❌")
                elif st.session_state.users[u2] != old_pwd:
                    st.error("Current password wrong ❌")
                elif new_pwd != cfm_pwd:
                    st.error("Passwords don't match ❌")
                elif len(new_pwd) < 4:
                    st.error("Min 4 characters")
                else:
                    st.session_state.users[u2] = new_pwd
                    st.success("Password changed ✅ Please login")
    st.stop()

st_autorefresh(interval=300000)

# ----------------------------------
# 💾 PERSIST CC MAP TO DISK
# ----------------------------------
CC_FILE = "cc_list.json"

def load_cc():
    """Load cc_author_map from disk. Returns dict {email: display_name}."""
    try:
        with open(CC_FILE) as f:
            data = json.load(f)
            if not data:
                return {}
            if isinstance(data[0], list):
                return dict(data)
            if isinstance(data[0], str):
                return {e: "" for e in data}
            return {}
    except:
        return {}

def save_cc(cc_map: dict):
    """Save cc_author_map to disk as list of [email, name] pairs."""
    try:
        with open(CC_FILE, "w") as f:
            json.dump(list(cc_map.items()), f)
    except:
        pass

# ----------------------------------
# 🔹 TOP BAR
# ----------------------------------
t1, t2, t3 = st.columns([4, 4, 1])
with t1:
    st.markdown(f"#### 🚀 QATrackPro &nbsp;|&nbsp; 👤 {st.session_state.username}")
with t2:
    run_input = st.text_input(
        "", placeholder="Run IDs or Plan IDs e.g. 1073095, P:5082",
        label_visibility="collapsed"
    )
with t3:
    if st.button("Logout", use_container_width=True):
        st.session_state.clear()
        st.rerun()

if not run_input:
    st.info("👆 Enter Run IDs or Plan IDs above (prefix plans with **P:** e.g. `P:5082`)")
    st.stop()

# ----------------------------------
# 🔹 HELPERS
# ----------------------------------
def extract_tester(comment):
    if not comment:
        return "Unknown"
    match = re.search(r"(Tester|Tested by)\s*[-:]\s*([A-Za-z ]+)", comment, re.IGNORECASE)
    return match.group(2).strip() if match else "Unknown"

@st.cache_data(ttl=3600)
def get_status_map():
    try:
        url = f"{st.secrets['testrail_url']}/index.php?/api/v2/get_statuses"
        res = session.get(url, auth=(st.secrets["testrail_email"], st.secrets["testrail_api"]), verify=False, timeout=30)
        return {s["id"]: s["name"].strip().title() for s in res.json()}
    except:
        return {}

@st.cache_data(ttl=600)
def get_runs_from_plan(plan_id):
    try:
        url = f"{st.secrets['testrail_url']}/index.php?/api/v2/get_plan/{plan_id}"
        res = session.get(url, auth=(st.secrets["testrail_email"], st.secrets["testrail_api"]), verify=False, timeout=30)
        if res.status_code != 200:
            st.error(f"❌ Plan {plan_id} failed: {res.status_code} {res.text[:150]}")
            return [], "", {}
        data      = res.json()
        plan_name = data.get("name", f"Plan {plan_id}")
        run_ids, run_names = [], {}
        for entry in data.get("entries", []):
            for run in entry.get("runs", []):
                rid = str(run["id"])
                run_ids.append(rid)
                run_names[rid] = run.get("name", f"Run {rid}")
        return run_ids, plan_name, run_names
    except Exception as e:
        st.error(f"❌ Plan {plan_id} error: {e}")
        return [], "", {}

@st.cache_data(ttl=600)
def get_testrail(run_id):
    try:
        url = f"{st.secrets['testrail_url']}/index.php?/api/v2/get_results_for_run/{run_id}"
        res = session.get(url, auth=(st.secrets["testrail_email"], st.secrets["testrail_api"]), verify=False, timeout=30)
        if res.status_code != 200:
            st.error(f"❌ Run {run_id} failed: {res.status_code} {res.text[:150]}")
            return pd.DataFrame()
        data = res.json()
    except Exception as e:
        st.error(f"❌ Run {run_id} error: {e}")
        return pd.DataFrame()

    results = data.get("results", []) if isinstance(data, dict) else data
    if not results:
        st.warning(f"⚠️ No results in Run ID {run_id}")
        return pd.DataFrame()

    status_map = get_status_map()
    allowed    = {"Passed", "Failed", "Blocked"}
    rows       = []

    for r in results:
        status = status_map.get(r.get("status_id"), "Unknown")
        ts = r.get("created_on")
        exec_date = datetime.fromtimestamp(ts).date() if ts else datetime.today().date()
        rows.append({
            "TestCaseID":    r.get("test_id"),
            "TesterName":    extract_tester(r.get("comment", "")),
            "Status":        status,
            "DefectID":      r.get("defects", ""),
            "ExecutionDate": exec_date,
            "RunID":         run_id,
            "created_on":    r.get("created_on", 0),
        })

    if not rows:
        return pd.DataFrame()

    raw_df = pd.DataFrame(rows)
    raw_df = raw_df.sort_values("created_on", ascending=False)
    raw_df = raw_df.drop_duplicates(subset=["TestCaseID"], keep="first")
    final  = raw_df[raw_df["Status"].isin(allowed)].copy()
    final  = final.drop(columns=["created_on"])
    return final

@st.cache_data(ttl=1800)
def get_jira(linked_bugs=()):
    domain = st.secrets["jira_domain"].strip().rstrip("/")
    if not domain.startswith("http"):
        domain = f"https://{domain}"
    keys = [k.strip() for k in st.secrets['jira_project_key'].split(",")]
    jql  = "project in (" + ", ".join(f'"{k}"' for k in keys) + ') AND updated >= "-3d" AND issuetype in ("Bug", "Story") ORDER BY updated DESC'
    issues = []
    start  = 0
    batch  = 1000
    while True:
        try:
            res = session.get(
                f"{domain}/rest/api/2/search",
                auth=HTTPBasicAuth(st.secrets["jira_email"], st.secrets["jira_token"]),
                params={"jql": jql, "maxResults": batch, "startAt": start, "fields": "key,status,updated,comment"},
                verify=False, timeout=30
            )
            if res.status_code != 200:
                st.error(f"❌ Jira API Error {res.status_code}: {res.text[:200]}")
                break
            data = res.json()
        except Exception as e:
            st.error(f"❌ Jira connection failed: {e}")
            break
        page = data.get("issues", [])
        if not page:
            break
        for issue in page:
            f             = issue["fields"]
            updated_str   = f.get("updated", "")[:10]
            comments_list = f.get("comment", {}).get("comments", [])
            if comments_list:
                last = comments_list[-1]
                body = last.get("body", "")
                try:
                    comment_txt = (
                        body["content"][0]["content"][0]["text"]
                        if isinstance(body, dict) else str(body).strip() or ""
                    )
                except:
                    comment_txt = ""
                comment_author = last["author"]["displayName"]
                comment_date   = datetime.strptime(last["created"][:10], "%Y-%m-%d")
                comment_id     = last.get("id", "")
            else:
                comment_txt = comment_author = comment_id = ""
                comment_date = None
            issues.append({
                "BugID":         issue["key"],
                "BugStatus":     f["status"]["name"],
                "UpdatedDate":   updated_str,
                "LatestComment": comment_txt or "No Comment",
                "Author":        comment_author,
                "CommentDate":   comment_date,
                "LastCommentID": comment_id,
            })
        start += batch
        if start >= data.get("total", 0):
            break
    return pd.DataFrame(issues) if issues else pd.DataFrame()


# ----------------------------------
# 🔹 GET LATEST COMMENT + @MENTIONS
# ----------------------------------
@st.cache_data(ttl=600)
def get_single_comment(bug_id: str):
    """
    Fetches the latest Jira comment and extracts @mentions in all formats:

      ADF (Atlassian Document Format):
        - 'mention' nodes carry the display name in attrs.text
          e.g. {"type":"mention","attrs":{"text":"Keerthan Kumar K",...}}

      Plain text / Wiki markup:
        - @email@domain.com  → captured as full email in MentionEmails
        - [~username]        → captured as display name in Mentions
        - @FirstName         → captured as display name in Mentions (emails excluded)

    Returns dict keys:
      Mentions       – list of display names / wiki usernames
      MentionEmails  – list of full email strings found via @email pattern
    """
    domain = st.secrets["jira_domain"].strip().rstrip("/")
    if not domain.startswith("http"):
        domain = f"https://{domain}"
    try:
        res = session.get(
            f"{domain}/rest/api/2/issue/{bug_id}/comment",
            auth=HTTPBasicAuth(st.secrets["jira_email"], st.secrets["jira_token"]),
            verify=False, timeout=20
        )
        if res.status_code != 200:
            return None
        data = res.json()
        if not data.get("comments"):
            return None

        last = data["comments"][-1]
        body = last.get("body", "")

        mentioned_names  = []   # display names from ADF / wiki / plain @Name
        mentioned_emails = []   # full emails like keerthan.kumar@hp.com
        txt = "No Comment"

        if isinstance(body, dict):
            # ── Atlassian Document Format (ADF) ──
            plain_parts = []

            def walk(node):
                if not isinstance(node, dict):
                    return
                ntype = node.get("type", "")
                if ntype == "mention":
                    # attrs.text contains the display name, e.g. "Keerthan Kumar K"
                    name = node.get("attrs", {}).get("text", "").lstrip("@").strip()
                    if name and name not in mentioned_names:
                        mentioned_names.append(name)
                elif ntype == "text":
                    plain_parts.append(node.get("text", ""))
                for child in node.get("content", []):
                    walk(child)

            walk(body)
            txt = " ".join(plain_parts).strip() or "No Comment"

        else:
            # ── Plain text / Wiki markup ──
            txt = str(body).strip() or "No Comment"

            # Pattern 1: @word@domain.com — full email mention
            for em in re.findall(r"@([\w.+-]+@[\w.-]+\.[a-zA-Z]{2,})", txt):
                if em not in mentioned_emails:
                    mentioned_emails.append(em)

            # Pattern 2: [~username] or [~accountid:xxx] — Jira wiki mention
            for wm in re.findall(r"\[~(?:accountid:)?([^\]]+)\]", txt):
                if wm not in mentioned_names:
                    mentioned_names.append(wm)

            # Pattern 3: plain @Word — exclude anything that is or leads to an email
            clean_txt = re.sub(r"@[\w.+-]+@[\w.-]+\.[a-zA-Z]{2,}", "", txt)
            for pm in re.findall(r"@([A-Za-z][A-Za-z0-9_]*(?:\s[A-Za-z][A-Za-z0-9_]*)*)", clean_txt):
                pm = pm.strip()
                if pm and pm not in mentioned_names:
                    mentioned_names.append(pm)

        dt = datetime.strptime(last["created"][:10], "%Y-%m-%d")
        return {
            "BugID":           bug_id,
            "LatestComment":   txt,
            "Author":          last["author"]["displayName"],
            "Mentions":        mentioned_names,
            "MentionEmails":   mentioned_emails,
            "CommentDateDisp": dt.strftime("%d %b %Y"),
            "CommentID":       last.get("id", ""),
        }
    except:
        return None

def resolve_mention_to_emails(mentions, mention_emails, cc_author_map):
    """
    KEY FIX: Jira strips the leading @ so full emails like
    'keerthan.kumar-k@hp.com' land in `mentions` (display names)
    instead of `mention_emails`. We detect and move them here.
    """
    EMAIL_RE    = re.compile(r"^[\w.\-+]+@[\w.\-]+\.[a-zA-Z]{2,}$")
    all_emails  = list(mention_emails)
    plain_names = []

    for m in mentions:
        if EMAIL_RE.match(m.strip()):
            if m.strip() not in all_emails:
                all_emails.append(m.strip())   # ✅ move email to email bucket
        else:
            plain_names.append(m)              # keep as display name

    to_list      = []
    matched_info = []

    for cc_email, mapped_name in cc_author_map.items():
        matched      = False
        match_reason = ""

        # Strategy 1: display name substring match
        for mention in plain_names:
            m = mention.lower().strip()
            n = mapped_name.lower().strip()
            if not m or not n:
                continue
            if m in n or n in m:
                matched      = True
                match_reason = f"@{mention} (name match) → {cc_email}"
                break

        # Strategy 2: exact email match
        # keerthan.kumar-k@hp.com == keerthan.kumar-k@hp.com ✅
        if not matched:
            for memail in all_emails:
                if memail.lower().strip() == cc_email.lower().strip():
                    matched      = True
                    match_reason = f"@{memail} (exact email) → {cc_email}"
                    break

        # Strategy 3: username token fuzzy match
        if not matched:
            name_tokens = set(re.split(r"[\s._\-]+", mapped_name.lower()))
            name_tokens.discard("")
            for memail in all_emails:
                username       = memail.split("@")[0].lower()
                username_parts = set(re.split(r"[\s._\-]+", username))
                username_parts.discard("")
                overlap        = name_tokens & username_parts

                if len(overlap) >= 2:
                    matched      = True
                    match_reason = f"@{memail} (tokens {overlap} ~ '{mapped_name}') → {cc_email}"
                    break
                elif len(overlap) == 1 and any(len(t) > 5 for t in overlap):
                    matched      = True
                    match_reason = f"@{memail} (strong token '{list(overlap)[0]}') → {cc_email}"
                    break
                else:
                    normalised = re.sub(r"[\.\-_]", " ", username)
                    if normalised in mapped_name.lower():
                        matched      = True
                        match_reason = f"@{memail} ('{normalised}' in '{mapped_name}') → {cc_email}"
                        break

        if matched and cc_email not in to_list:
            to_list.append(cc_email)
            matched_info.append(match_reason)

    return to_list, matched_info


# ----------------------------------
# 🔹 SENDGRID EMAIL NOTIFICATION
# ----------------------------------
def send_via_sendgrid(to_list, bug_id, comment_txt, author, comment_date):
    """Send branded @mention notification via SendGrid API."""
    if not to_list:
        return False, "No recipients"

    domain = st.secrets["jira_domain"].strip().rstrip("/")
    if not domain.startswith("http"):
        domain = f"https://{domain}"
    jira_url = f"{domain}/browse/{bug_id}"

    html = f"""
    <html><body style="font-family:Arial,sans-serif;background:#f4f4f4;padding:20px;">
    <div style="max-width:600px;margin:auto;background:white;border-radius:10px;overflow:hidden;">
        <div style="background:#1F2A40;padding:20px;">
            <h2 style="color:white;margin:0;">🚀 QATrackPro — You were @mentioned!</h2>
        </div>
        <div style="padding:24px;">
            <table style="width:100%;border-collapse:collapse;">
                <tr>
                    <td style="padding:8px;color:#666;width:140px;">Bug ID</td>
                    <td style="padding:8px;font-weight:bold;">
                        <a href="{jira_url}" style="color:#3498DB;">{bug_id}</a>
                    </td>
                </tr>
                <tr style="background:#f9f9f9;">
                    <td style="padding:8px;color:#666;">Comment By</td>
                    <td style="padding:8px;">{author}</td>
                </tr>
                <tr>
                    <td style="padding:8px;color:#666;">Date</td>
                    <td style="padding:8px;">{comment_date}</td>
                </tr>
            </table>
            <div style="margin-top:16px;padding:16px;background:#f0f7ff;
                        border-left:4px solid #3498DB;border-radius:4px;">
                <p style="margin:0;font-size:14px;color:#333;">{comment_txt}</p>
            </div>
            <div style="margin-top:20px;text-align:center;">
                <a href="{jira_url}" style="background:#3498DB;color:white;padding:10px 24px;
                   border-radius:6px;text-decoration:none;font-weight:bold;">
                   View in Jira →
                </a>
            </div>
            <p style="margin-top:24px;font-size:11px;color:#999;text-align:center;">
                Sent by QATrackPro Dashboard
            </p>
        </div>
    </div>
    </body></html>
    """

    payload = {
        "personalizations": [{"to": [{"email": e} for e in to_list]}],
        "from":    {"email": st.secrets["sendgrid_sender"], "name": "QATrackPro"},
        "subject": f"[QATrackPro] You were @mentioned on {bug_id}",
        "content": [{"type": "text/html", "value": html}]
    }

    try:
        res = requests.post(
            "https://api.sendgrid.com/v3/mail/send",
            headers={
                "Authorization": f"Bearer {st.secrets['sendgrid_api_key']}",
                "Content-Type":  "application/json"
            },
            json=payload,
            timeout=20
        )
        if res.status_code == 202:
            return True, f"Sent to {len(to_list)} recipients"
        return False, f"SendGrid {res.status_code}: {res.text[:100]}"
    except Exception as e:
        return False, str(e)


# ----------------------------------
# 🔹 RESOLVE PLAN IDs → RUN IDs
# ----------------------------------
raw_ids      = [r.strip() for r in run_input.split(",") if r.strip()]
all_run_ids  = []
run_name_map = {}
plan_info    = {}

with st.spinner("Resolving IDs..."):
    for id_ in raw_ids:
        if id_.upper().startswith("P:"):
            plan_id = id_[2:].strip()
            child_runs, plan_name, run_names = get_runs_from_plan(plan_id)
            plan_info[plan_id] = {"name": plan_name, "run_ids": child_runs}
            for rid in child_runs:
                all_run_ids.append(rid)
                run_name_map[rid] = run_names.get(rid, f"Run {rid}")
        else:
            all_run_ids.append(id_)
            run_name_map[id_] = f"Run {id_}"

all_run_ids = list(dict.fromkeys(all_run_ids))

if plan_info:
    for pid, info in plan_info.items():
        st.markdown(
            f'<span class="plan-badge">📋 {info["name"]} (P:{pid}) → {len(info["run_ids"])} runs</span>',
            unsafe_allow_html=True
        )

if st.session_state.get("last_input") != run_input:
    st.session_state.last_input = run_input
    for rid in all_run_ids:
        st.session_state[f"chk_{rid}"] = True

selected_count = sum(1 for rid in all_run_ids if st.session_state.get(f"chk_{rid}", True))

with st.expander(f"▶ Select Runs — {selected_count} of {len(all_run_ids)} selected", expanded=False):
    sa1, sa2, _ = st.columns([1, 1, 4])
    with sa1:
        if st.button("✅ Select All", use_container_width=True):
            for rid in all_run_ids:
                st.session_state[f"chk_{rid}"] = True
            st.rerun()
    with sa2:
        if st.button("❌ Clear All", use_container_width=True):
            for rid in all_run_ids:
                st.session_state[f"chk_{rid}"] = False
            st.rerun()
    st.markdown("---")
    cols_per_row = 2
    rows_list = [all_run_ids[i:i+cols_per_row] for i in range(0, len(all_run_ids), cols_per_row)]
    for row_runs in rows_list:
        cols = st.columns(cols_per_row)
        for i, rid in enumerate(row_runs):
            with cols[i]:
                run_name = run_name_map.get(rid, f"Run {rid}")
                st.checkbox(f"**{rid}**  \n{run_name}", key=f"chk_{rid}")

run_ids = [rid for rid in all_run_ids if st.session_state.get(f"chk_{rid}", True)]

if not run_ids:
    st.warning("⚠️ Please select at least one run")
    st.stop()

st.caption(f"📊 Using **{len(run_ids)}** of {len(all_run_ids)} runs")

# ----------------------------------
# 🔹 LOAD DATA
# ----------------------------------
jira_result = {}

def _load_jira(linked):
    try:
        jira_result["df"] = get_jira(linked)
    except:
        jira_result["df"] = pd.DataFrame()

jira_thread = threading.Thread(target=_load_jira, args=((),))
jira_thread.start()

with st.spinner("Loading TestRail data..."):
    df_list = [get_testrail(r) for r in run_ids]
    df_list = [d for d in df_list if not d.empty]

if not df_list:
    st.error("❌ No data found for selected runs")
    jira_thread.join()
    st.stop()

test_df     = pd.concat(df_list, ignore_index=True)
linked_bugs = tuple(test_df["DefectID"].replace("", pd.NA).dropna().unique().tolist())

jira_thread.join()

with st.spinner("Loading Jira data..."):
    jira_df = get_jira(linked_bugs)

if jira_df is None or jira_df.empty or "BugID" not in jira_df.columns:
    st.warning("⚠️ Jira not available — showing TestRail data only")
    final_df = test_df.copy()
    for col in ["BugID", "BugStatus", "LatestComment", "Author", "CommentDate"]:
        final_df[col] = "N/A"
else:
    final_df = pd.merge(test_df, jira_df, left_on="DefectID", right_on="BugID", how="left")

status_colors = {"Passed": "#2ECC71", "Failed": "#E74C3C", "Blocked": "#F39C12"}

# ----------------------------------
# 🔹 TESTER FILTER + KPI
# ----------------------------------
fa, _ = st.columns([3, 5])
with fa:
    all_testers     = sorted(final_df["TesterName"].dropna().unique().tolist())
    selected_tester = st.selectbox("👤 Filter by Tester", options=["All"] + all_testers)

filtered_df = final_df.copy() if selected_tester == "All" else final_df[final_df["TesterName"] == selected_tester].copy()

total   = len(filtered_df)
passed  = len(filtered_df[filtered_df["Status"] == "Passed"])
failed  = len(filtered_df[filtered_df["Status"] == "Failed"])
blocked = len(filtered_df[filtered_df["Status"] == "Blocked"])
rate    = round((passed / total) * 100, 2) if total else 0
bugs    = filtered_df["DefectID"].replace("", pd.NA).dropna().nunique()

c1, c2, c3, c4, c5, c6 = st.columns(6)
c1.metric("Total",   total)
c2.metric("Passed",  passed)
c3.metric("Failed",  failed)
c4.metric("Blocked", blocked)
c5.metric("Pass %",  f"{rate}%")
c6.metric("Bugs",    bugs)

st.markdown("---")

# ----------------------------------
# 🔲 5 NAV CARDS
# ----------------------------------
if "active_section" not in st.session_state:
    st.session_state.active_section = None

cards = [
    ("📈", "Execution Trend",    "trend"),
    ("👨‍💻", "Tester Performance", "perf"),
    ("🚨", "Failed Tests",       "failed"),
    ("💬", "Jira Activity",      "jira"),
    ("📊", "Internal MPPM",      "npm"),
]

nav1, nav2, nav3, nav4, nav5 = st.columns(5)
for col, (icon, label, key) in zip([nav1, nav2, nav3, nav4, nav5], cards):
    with col:
        is_active = st.session_state.active_section == key
        border    = "#3498DB" if is_active else "#2C3E50"
        bg        = "#1a3a5c" if is_active else "#1F2A40"
        lbl_color = "#3498DB" if is_active else "#BDC3C7"
        st.markdown(f"""
        <div style="background:{bg}; border:2px solid {border}; border-radius:14px;
                    padding:20px 12px; text-align:center; margin-bottom:4px;">
            <div style="font-size:30px;">{icon}</div>
            <div style="font-size:13px; font-weight:600; color:{lbl_color}; margin-top:6px;">{label}</div>
        </div>
        """, unsafe_allow_html=True)
        if st.button(f"{'✓ ' if is_active else ''}{label}", key=f"btn_{key}", use_container_width=True):
            st.session_state.active_section = None if st.session_state.active_section == key else key
            st.rerun()

st.markdown("---")

# ----------------------------------
# 📈 EXECUTION TREND
# ----------------------------------
if st.session_state.active_section == "trend":
    st.markdown('<div class="section-header">📈 Execution Trend</div>', unsafe_allow_html=True)
    filtered_df["ExecutionDate"] = pd.to_datetime(filtered_df["ExecutionDate"])
    tc1, tc2 = st.columns(2)
    with tc1:
        start_date = st.date_input("🗓️ Run Start Date", value=filtered_df["ExecutionDate"].min().date())
    with tc2:
        end_date = st.date_input("🗓️ Run End Date", value=filtered_df["ExecutionDate"].max().date())

    start_ts    = pd.Timestamp(start_date)
    end_ts      = pd.Timestamp(end_date)
    total_tests = len(filtered_df)
    daily = filtered_df.groupby(filtered_df["ExecutionDate"].dt.date).size().reset_index(name="DailyCount")
    daily["ExecutionDate"] = pd.to_datetime(daily["ExecutionDate"])
    all_dates = pd.date_range(start=start_ts, end=end_ts, freq="D")
    date_df   = pd.DataFrame({"Date": all_dates})
    date_df   = pd.merge(date_df, daily.rename(columns={"ExecutionDate": "Date"}), on="Date", how="left")
    date_df["DailyCount"] = date_df["DailyCount"].fillna(0)
    date_df["Actual"]     = date_df["DailyCount"].cumsum()
    n_days = max((end_ts - start_ts).days, 1)
    date_df["Ideal"] = [round(total_tests * i / n_days) for i in range(len(date_df))]
    last_actual  = date_df["Actual"].iloc[-1]
    last_ideal   = date_df["Ideal"].iloc[-1]
    on_track     = last_actual >= last_ideal
    behind       = int(last_ideal - last_actual)
    track_label  = "✅ On Track" if on_track else f"⚠️ Off Track — {behind} tests behind"
    track_color  = "#2ECC71" if on_track else "#E74C3C"
    st.markdown(f"<span style='background:{track_color}22; color:{track_color}; padding:5px 14px; border-radius:8px; font-weight:600; font-size:15px;'>{track_label}</span>", unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    fig_burn = px.line(date_df, x="Date", y=["Ideal", "Actual"], markers=True,
                       color_discrete_map={"Ideal": "#95A5A6", "Actual": track_color},
                       labels={"value": "Tests Executed", "variable": ""})
    fig_burn.update_layout(plot_bgcolor="#1F2A40", paper_bgcolor="#1F2A40", font_color="white",
                           xaxis=dict(showgrid=False), yaxis=dict(showgrid=True, gridcolor="#2C3E50", rangemode="tozero"),
                           legend=dict(orientation="h", y=1.1))
    st.plotly_chart(fig_burn, use_container_width=True)
    if "RunID" in filtered_df.columns and len(run_ids) > 1:
        st.markdown("#### 📊 Per-Run Breakdown")
        run_summary = filtered_df.groupby(["RunID", "Status"]).size().reset_index(name="Count")
        run_summary["RunLabel"] = run_summary["RunID"].apply(lambda x: f"{x} — {run_name_map.get(x, '')}")
        fig_run = px.bar(run_summary, x="RunLabel", y="Count", color="Status", barmode="group", text="Count",
                         color_discrete_map=status_colors, labels={"RunLabel": "Run", "Count": "Tests"})
        fig_run.update_layout(plot_bgcolor="#1F2A40", paper_bgcolor="#1F2A40", font_color="white",
                              xaxis=dict(showgrid=False, type="category"), yaxis=dict(showgrid=True, gridcolor="#2C3E50"))
        st.plotly_chart(fig_run, use_container_width=True)

# ----------------------------------
# 👨‍💻 TESTER PERFORMANCE
# ----------------------------------
elif st.session_state.active_section == "perf":
    st.markdown('<div class="section-header">👨‍💻 Tester Performance</div>', unsafe_allow_html=True)
    filtered_df["ExecutionDate"] = pd.to_datetime(filtered_df["ExecutionDate"])
    tab1, tab2, tab3 = st.tabs(["📅 Daily", "📆 Weekly", "🗂️ All"])

    def make_perf_chart(df, title):
        p = df.groupby(["TesterName", "Status"]).size().reset_index(name="TestCount")
        fig = px.bar(p, x="TesterName", y="TestCount", color="Status", barmode="group",
                     text="TestCount", color_discrete_map=status_colors,
                     title=title, labels={"TestCount": "Tests", "TesterName": "Tester"})
        fig.update_layout(plot_bgcolor="#1F2A40", paper_bgcolor="#1F2A40", font_color="white",
                          xaxis=dict(showgrid=False), yaxis=dict(showgrid=True, gridcolor="#2C3E50"))
        return fig

    with tab1:
        today    = pd.Timestamp(datetime.today().date())
        daily_df = filtered_df[filtered_df["ExecutionDate"] == today]
        if daily_df.empty:
            st.info("No tests executed today yet.")
        else:
            st.plotly_chart(make_perf_chart(daily_df, f"Daily — {today.strftime('%d %b %Y')}"), use_container_width=True)
            st.dataframe(daily_df.groupby("TesterName")["Status"].value_counts().unstack(fill_value=0).reset_index(), use_container_width=True)

    with tab2:
        week_start = pd.Timestamp(datetime.today().date() - timedelta(days=7))
        weekly_df  = filtered_df[filtered_df["ExecutionDate"] >= week_start]
        if weekly_df.empty:
            st.info("No tests in the last 7 days.")
        else:
            st.plotly_chart(make_perf_chart(weekly_df, "Weekly — last 7 days"), use_container_width=True)
            st.dataframe(weekly_df.groupby("TesterName")["Status"].value_counts().unstack(fill_value=0).reset_index(), use_container_width=True)

    with tab3:
        st.plotly_chart(make_perf_chart(filtered_df, "All Runs"), use_container_width=True)

    st.markdown("---")

    def generate_pdf(df, title="Report"):
        file  = f"{title}.pdf"
        doc   = SimpleDocTemplate(file, pagesize=letter)
        data  = [df.columns.tolist()] + df.astype(str).values.tolist()
        table = Table(data)
        style = TableStyle([
            ("BACKGROUND",    (0,0),(-1,0), rl_colors.grey),
            ("TEXTCOLOR",     (0,0),(-1,0), rl_colors.whitesmoke),
            ("ALIGN",         (0,0),(-1,-1),"CENTER"),
            ("FONTNAME",      (0,0),(-1,0), "Helvetica-Bold"),
            ("BOTTOMPADDING", (0,0),(-1,0), 12),
            ("GRID",          (0,0),(-1,-1),0.5, rl_colors.black),
        ])
        table.setStyle(style)
        doc.build([table])
        return file

    def tester_summary(df):
        df = df.copy()
        df["ExecutionDate"] = pd.to_datetime(df["ExecutionDate"])
        summary = df.groupby(["TesterName", "Status"]).size().unstack(fill_value=0).reset_index()
        for col in ["Passed", "Failed", "Blocked"]:
            if col not in summary.columns:
                summary[col] = 0
        summary["Total"]  = summary["Passed"] + summary["Failed"] + summary["Blocked"]
        summary["Pass %"] = (summary["Passed"] / summary["Total"] * 100).round(1).astype(str) + "%"
        return summary[["TesterName"] + [c for c in ["Passed","Failed","Blocked","Total","Pass %"] if c in summary.columns]]

    pr1, pr2 = st.columns(2)
    with pr1:
        if st.button("📄 Daily PDF", use_container_width=True):
            today  = pd.Timestamp(datetime.today().date())
            day_df = filtered_df[pd.to_datetime(filtered_df["ExecutionDate"]) == today]
            summ   = tester_summary(day_df) if not day_df.empty else pd.DataFrame({"Note": ["No tests today"]})
            pdf    = generate_pdf(summ, "daily_report")
            with open(pdf, "rb") as f:
                st.download_button("⬇️ Download Daily", f, file_name="daily_tester_report.pdf", use_container_width=True)
    with pr2:
        if st.button("📄 Weekly PDF", use_container_width=True):
            wk    = pd.Timestamp(datetime.today().date() - timedelta(days=7))
            wk_df = filtered_df[pd.to_datetime(filtered_df["ExecutionDate"]) >= wk]
            summ  = tester_summary(wk_df) if not wk_df.empty else pd.DataFrame({"Note": ["No tests this week"]})
            pdf   = generate_pdf(summ, "weekly_report")
            with open(pdf, "rb") as f:
                st.download_button("⬇️ Download Weekly", f, file_name="weekly_tester_report.pdf", use_container_width=True)

# ----------------------------------
# 🚨 FAILED TESTS
# ----------------------------------
elif st.session_state.active_section == "failed":
    st.markdown('<div class="section-header">🚨 Failed Tests</div>', unsafe_allow_html=True)
    failed_df = filtered_df[filtered_df["Status"] == "Failed"]
    if failed_df.empty:
        st.success("✅ No failed tests!")
    else:
        cols = ["TestCaseID", "TesterName", "Status", "DefectID"]
        if "RunID" in failed_df.columns:
            cols.insert(0, "RunID")
        if "BugStatus" in failed_df.columns:
            cols.append("BugStatus")
        st.dataframe(failed_df[cols], use_container_width=True)
        if "RunID" in failed_df.columns and len(run_ids) > 1:
            st.markdown("#### 📊 Failed Tests by Run")
            failed_run = failed_df.groupby("RunID").size().reset_index(name="FailedCount")
            failed_run["RunLabel"] = failed_run["RunID"].apply(lambda x: f"{x} — {run_name_map.get(x, '')}")
            fig_fr = px.bar(failed_run, x="RunLabel", y="FailedCount", text="FailedCount",
                            color_discrete_sequence=["#E74C3C"],
                            labels={"RunLabel": "Run", "FailedCount": "Failed Tests"})
            fig_fr.update_layout(plot_bgcolor="#1F2A40", paper_bgcolor="#1F2A40", font_color="white",
                                 xaxis=dict(showgrid=False, type="category"), yaxis=dict(showgrid=True, gridcolor="#2C3E50"))
            st.plotly_chart(fig_fr, use_container_width=True)

# ----------------------------------
# 💬 JIRA ACTIVITY + @MENTION NOTIFICATIONS
# ----------------------------------
elif st.session_state.active_section == "jira":
    st.markdown('<div class="section-header">💬 Jira Activity</div>', unsafe_allow_html=True)

    if jira_df is None or jira_df.empty:
        st.info("Jira not connected")
    else:
        jira_df["Project"] = jira_df["BugID"].str.extract(r"^([A-Z]+)-\d+")
        jira_projects      = sorted(jira_df["Project"].dropna().unique().tolist())
        selected_jira_project = st.selectbox(
            "🔍 Filter by Project", options=["All"] + jira_projects, key="jira_project_filter"
        )

        filtered_jira = (
            jira_df if selected_jira_project == "All"
            else jira_df[jira_df["Project"] == selected_jira_project]
        ).copy()

        filtered_jira["CommentDate"] = pd.to_datetime(filtered_jira["CommentDate"], errors="coerce")
        filtered_jira = filtered_jira.sort_values("CommentDate", ascending=False)
        filtered_jira["CommentDate"] = filtered_jira["CommentDate"].dt.strftime("%d %b %Y")

        display_cols = ["BugID", "BugStatus", "LatestComment", "Author", "CommentDate"]
        existing     = [c for c in display_cols if c in filtered_jira.columns]
        st.dataframe(filtered_jira[existing], use_container_width=True)

        # ======================================================
        # 🔔 @MENTION NOTIFICATIONS via SendGrid
        # ======================================================
        st.markdown("---")
        st.markdown("### 🔔 @Mention Notifications")
        st.caption(
            "When someone adds a Jira comment that **@mentions** a person in your CC list, "
            "an email is sent **only to that mentioned person**. Checks every 5 min automatically."
        )
 
        # ── init session state ────────────────────────────────
        if "seen_comment_ids" not in st.session_state:
            st.session_state.seen_comment_ids = {}
        if "cc_author_map" not in st.session_state:
            st.session_state.cc_author_map = load_cc()
        if "notif_log" not in st.session_state:
            st.session_state.notif_log = []
 
        # ── CC list UI ────────────────────────────────────────
        st.markdown("#### 👥 CC List — Email → Jira Display Name")
        st.caption(
            "Enter each person's **email** and their **exact Jira display name** "
            "(as it appears when you @mention them in a Jira comment)."
        )
 
        map_col1, map_col2, map_col3 = st.columns([3, 3, 1])
        with map_col1:
            new_email = st.text_input(
                "📧 Email", placeholder="keerthan.kumar-k@hp.com", key="new_map_email"
            )
        with map_col2:
            new_author = st.text_input(
                "👤 Jira Display Name (exact)", placeholder="Keerthan Kumar K", key="new_map_author"
            )
        with map_col3:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("➕ Add", use_container_width=True, key="add_cc_btn"):
                if new_email and "@" in new_email and new_author.strip():
                    st.session_state.cc_author_map[new_email.strip()] = new_author.strip()
                    save_cc(st.session_state.cc_author_map)
                    st.success(f"Added: {new_email.strip()} → {new_author.strip()} ✅")
                    st.rerun()
                else:
                    st.warning("Please enter a valid email and display name.")
 
        if st.session_state.cc_author_map:
            st.markdown("**Current CC List:**")
            for em, au in list(st.session_state.cc_author_map.items()):
                rc1, rc2 = st.columns([8, 1])
                with rc1:
                    st.markdown(
                        f'<span class="notif-badge">📧 {em}</span>'
                        f'&nbsp;→&nbsp;<span style="color:#BDC3C7;font-size:13px;">@{au}</span>',
                        unsafe_allow_html=True
                    )
                with rc2:
                    if st.button("🗑️", key=f"del_{em}"):
                        del st.session_state.cc_author_map[em]
                        save_cc(st.session_state.cc_author_map)
                        st.rerun()
            st.markdown("<br>", unsafe_allow_html=True)
        else:
            st.info("No CC entries yet — add emails above.")
 
        # ── toggle ───────────────────────────────────────────
        auto_on = st.toggle(
            "🔄 Auto-notify on @mentions (checks every 5 min)",
            value=st.session_state.get("auto_notify_on", False),
            key="auto_notify_toggle"
        )
        st.session_state.auto_notify_on = auto_on
 
        # ── AUTO CHECK + SEND ─────────────────────────────────
        if auto_on and linked_bugs:
            if not st.session_state.cc_author_map:
                st.warning("⚠️ Add at least one email to the CC list above.")
            else:
                newly_notified = []
 
                for bug_id in linked_bugs:
                    try:
                        latest = get_single_comment(bug_id)
                        if not latest:
                            continue
 
                        current_id = latest.get("CommentID", "")
                        last_seen  = st.session_state.seen_comment_ids.get(bug_id, "")
 
                        if not current_id or current_id == last_seen:
                            continue
 
                        # Mark seen immediately
                        st.session_state.seen_comment_ids[bug_id] = current_id
 
                        author         = latest.get("Author", "")
                        mentions       = latest.get("Mentions", [])
                        mention_emails = latest.get("MentionEmails", [])
 
                        all_mentions_display = (
                            [f"@{m}" for m in mentions] +
                            [f"@{e}" for e in mention_emails]
                        )
 
                        if not mentions and not mention_emails:
                            st.session_state.notif_log.append({
                                "time":           datetime.now().strftime("%d %b %Y %H:%M"),
                                "bug":            bug_id,
                                "comment_author": author,
                                "mentioned":      "— (no @mentions)",
                                "matched":        "—",
                                "sent_to":        0,
                                "emails":         "—",
                                "status":         "⚠️ No @mentions in comment",
                            })
                            continue
 
                        to_list, matched_info = resolve_mention_to_emails(
                            mentions, mention_emails, st.session_state.cc_author_map
                        )

                        if to_list:
                            ok, msg = send_via_sendgrid(
                                to_list,
                                bug_id,
                                latest.get("LatestComment", ""),
                                author,
                                latest.get("CommentDateDisp", ""),
                            )
                            st.session_state.notif_log.append({
                                "time":           datetime.now().strftime("%d %b %Y %H:%M"),
                                "bug":            bug_id,
                                "comment_author": author,
                                "mentioned":      ", ".join(all_mentions_display),
                                "matched":        " | ".join(matched_info),
                                "sent_to":        len(to_list),
                                "emails":         ", ".join(to_list),
                                "status":         "✅ Sent" if ok else f"❌ {msg[:60]}",
                            })
                            newly_notified.append(bug_id)
                        else:
                            st.session_state.notif_log.append({
                                "time":           datetime.now().strftime("%d %b %Y %H:%M"),
                                "bug":            bug_id,
                                "comment_author": author,
                                "mentioned":      ", ".join(all_mentions_display),
                                "matched":        "—",
                                "sent_to":        0,
                                "emails":         "—",
                                "status":         "⚠️ @mentioned person not in CC list",
                            })
 
                    except:
                        continue
 
                if newly_notified:
                    st.success(f"🔔 Notifications sent for: {', '.join(newly_notified)}")
                else:
                    st.caption(f"✅ Checked for @mentions — last checked {datetime.now().strftime('%H:%M')}")
 
        elif auto_on and not linked_bugs:
            st.info("No linked bugs found in current runs.")
 
        # ── Notification log ──────────────────────────────────
        if st.session_state.notif_log:
            st.markdown("#### 📋 Notification Log")
            log_df = pd.DataFrame(st.session_state.notif_log[::-1])
            st.dataframe(log_df, use_container_width=True)
            if st.button("🗑️ Clear Log"):
                st.session_state.notif_log = []
                st.rerun()
 

# ----------------------------------
# 📊 NPM — PROJECT METRICS EXPORT
# ----------------------------------
elif st.session_state.active_section == "npm":
    st.markdown('<div class="section-header">📊 Internal MPPM — Project Metrics Export</div>', unsafe_allow_html=True)

    @st.cache_data(ttl=600)
    def get_run_case_count(run_id):
        try:
            url = (
                f"{st.secrets['testrail_url']}/index.php?/api/v2/get_tests/{run_id}"
                f"&limit=250&offset=0"
            )
            res = session.get(
                url,
                auth=(st.secrets["testrail_email"], st.secrets["testrail_api"]),
                verify=False, timeout=30
            )
            if res.status_code != 200:
                return 0
            data  = res.json()
            tests = data.get("tests", data) if isinstance(data, dict) else data
            return len(tests)
        except:
            return 0

    total_manual = sum(get_run_case_count(r) for r in run_ids)

    failed_only = test_df[test_df["Status"] == "Failed"]
    linked_defect_ids = (
        failed_only["DefectID"]
        .replace("", pd.NA).dropna()
        .str.split(",").explode().str.strip().dropna()
        .unique().tolist()
    )
    total_defects   = len(linked_defect_ids)
    total_testcases = total_manual

    st.caption("Columns marked **auto** are filled from TestRail / Jira. Fill the rest after exporting.")

    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("Manual Testcases (TestRail)", total_manual)
    k2.metric("Total Testcases", total_testcases)
    k3.metric("Total Defects (Jira)", total_defects)
    k4.metric("Resources", "—")
    k5.metric("Working Days", "—")

    st.markdown("---")

    if "npm_rows" not in st.session_state:
        st.session_state.npm_rows = []

    if not st.session_state.npm_rows:
        st.session_state.npm_rows = [{
            "Ink Project": "—",
            "No. of Manual Testcases (TestRail Test Count)": total_manual,
            "Regression/Exploratory/Defect Validation/Sanity": "",
            "Adhoc": "",
            "Total Testcases": total_testcases,
            "Total No. of Defects Submitted from Manual": total_defects,
            "Number of Resources Utilized": "",
            "No. of Working Days": "",
            "Comments": "",
        }]

    with st.expander("➕ Add / Edit a Project Row", expanded=True):
        fc1, fc2, fc3 = st.columns(3)
        with fc1:
            np_project = st.text_input("Ink Project", key="np_project")
            np_reg_etc = st.number_input(
                "Regression / Exploratory / Defect Validation / Sanity",
                min_value=0, value=0, key="np_reg_etc"
            )
            np_adhoc = st.number_input("Adhoc", min_value=0, value=0, key="np_adhoc")
        with fc2:
            np_res  = st.number_input("Number of Resources Utilized", min_value=0, value=0, key="np_res")
            np_days = st.number_input("No. of Working Days", min_value=0, value=0, key="np_days")
        with fc3:
            np_comments    = st.text_area("Comments", key="np_comments", height=80)
            computed_total = np_reg_etc + np_adhoc
            st.markdown(f"**Manual Testcases (auto):** `{total_manual}`")
            st.markdown(f"**Total Defects (auto):** `{total_defects}`")
            st.markdown(f"**Total Testcases (sum):** `{computed_total or total_testcases}`")

        if st.button("➕ Add Row", type="primary"):
            st.session_state.npm_rows.append({
                "Ink Project": np_project or "—",
                "No. of Manual Testcases (TestRail Test Count)": total_manual,
                "Regression/Exploratory/Defect Validation/Sanity": np_reg_etc or "",
                "Adhoc": np_adhoc or "",
                "Total Testcases": computed_total or total_testcases,
                "Total No. of Defects Submitted from Manual": total_defects,
                "Number of Resources Utilized": np_res or "",
                "No. of Working Days": np_days or "",
                "Comments": np_comments or "",
            })
            st.rerun()

    if st.session_state.npm_rows:
        npm_df = pd.DataFrame(st.session_state.npm_rows)
        st.dataframe(npm_df, use_container_width=True)

        col_del, col_clr, _ = st.columns([1, 1, 4])
        with col_del:
            del_idx = st.number_input("Delete row #", min_value=1,
                                      max_value=len(st.session_state.npm_rows), value=1, step=1)
            if st.button("🗑️ Delete Row"):
                st.session_state.npm_rows.pop(del_idx - 1)
                st.rerun()
        with col_clr:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("❌ Clear All Rows"):
                st.session_state.npm_rows = []
                st.rerun()

        st.markdown("---")

        def build_npm_excel(rows):
            wb = Workbook()
            ws = wb.active
            ws.title = "NPM Metrics"

            headers     = list(rows[0].keys())
            header_fill = PatternFill("solid", fgColor="1F2A40")
            header_font = Font(bold=True, color="FFFFFF", size=11)

            for ci, h in enumerate(headers, 1):
                cell           = ws.cell(row=1, column=ci, value=h)
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

            auto_cols = {
                "No. of Manual Testcases (TestRail Test Count)",
                "Total Testcases",
                "Total No. of Defects Submitted from Manual",
            }
            auto_fill  = PatternFill("solid", fgColor="D4EDDA")
            empty_fill = PatternFill("solid", fgColor="FFF3CD")

            for ri, row in enumerate(rows, 2):
                for ci, h in enumerate(headers, 1):
                    val            = row[h]
                    cell           = ws.cell(row=ri, column=ci, value=val if val != "" else None)
                    cell.alignment = Alignment(horizontal="center")
                    if h in auto_cols:
                        cell.fill = auto_fill
                    elif val == "":
                        cell.fill = empty_fill

            col_widths = [20, 28, 38, 10, 16, 32, 24, 18, 20]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[ws.cell(1, i).column_letter].width = w

            ws.row_dimensions[1].height = 48
            ws.freeze_panes = "A2"

            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf

        excel_buf = build_npm_excel(st.session_state.npm_rows)
        st.download_button(
            label="📥 Download MPPM Excel",
            data=excel_buf,
            file_name="MPPM_Project_Metrics.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=False,
        )
        st.caption("Green cells = auto-filled from TestRail/Jira. Amber cells = fill these in Excel after downloading.")
